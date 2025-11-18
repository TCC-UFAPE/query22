import pandas as pd
from openpyxl import load_workbook

EXCEL_FILE = 'Relatorio_Analise_CVEs_LLM.xlsx'

# Listar abas e primeiras 5 linhas
wb = load_workbook(EXCEL_FILE, read_only=True)
print('Abas encontradas:', wb.sheetnames)

for sheet in wb.sheetnames:
    print('\n--- Sheet:', sheet)
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet, engine='openpyxl')
    print('Colunas:', list(df.columns))
    print(df.head(3))
