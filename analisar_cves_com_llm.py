import os
import json
from groq import Groq
import sys
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Configuração da API
try:
    with open("config.json", "r", encoding="utf-8") as config_file:
        config = json.load(config_file)
        api_key = config["api_key"]
except FileNotFoundError:
    print("Erro: Arquivo 'config.json' não encontrado. Crie este arquivo com sua chave da API Groq.")
    print("Exemplo de config.json: {\"api_key\": \"SUA_CHAVE_API_AQUI\"}")
    sys.exit(1)
except KeyError:
    print("Erro: Chave 'api_key' não encontrada no arquivo 'config.json'.")
    sys.exit(1)

client = Groq(api_key=api_key)

# Extensões de arquivos de código para analisar
CODE_FILE_EXTENSIONS = {
    '.cpp', '.c', '.h', '.hpp', '.java', '.py', '.js', '.ts', '.cc', '.conf', '.sym',
    '.html', '.css', '.go', '.rs', '.php', '.rb', '.swift', '.kt', '.xml', '.am', '.ac',
    '.sh', '.bash', '.yml', '.yaml', '.json', '.sql', '.pl', '.r', '.scala', '.m'
}

def remove_comments(code):
    """Remove comentários do código"""
    code = re.sub(r'/\*.*?\*/', '', code, flags=re.DOTALL)
    code = re.sub(r'(//|#).*', '', code)
    code = "\n".join(line.strip() for line in code.splitlines() if line.strip())
    return code


def read_all_code_files(commit_folder):
    """
    Lê todos os arquivos de código em uma pasta de commit
    Retorna um dicionário com {filename: content}
    """
    code_files = {}
    
    for root, dirs, files in os.walk(commit_folder):
        for filename in files:
            # Ignorar arquivos especiais
            if filename.endswith('.DELETED') or filename.endswith('.metadata') or filename == 'README.md':
                continue
            
            # Verificar se é um arquivo de código
            if any(filename.endswith(ext) for ext in CODE_FILE_EXTENSIONS):
                filepath = os.path.join(root, filename)
                relative_path = os.path.relpath(filepath, commit_folder)
                
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    
                    if content.strip():
                        code_files[relative_path] = content
                except Exception as e:
                    print(f"      ⚠ Erro ao ler {relative_path}: {e}")
    
    return code_files


def analyze_commit_with_llm(cve_name, commit_hash, code_files, model_name):
    """
    Analisa todos os arquivos de um commit usando a LLM
    """
    if not code_files:
        return {
            "vulnerability": "N/A",
            "vulnerability_type": "N/A",
            "vulnerability_name": "N/A",
            "explanation": "Nenhum arquivo de código encontrado para análise"
        }
    
    # Construir o código completo para análise
    combined_code = ""
    for filename, content in code_files.items():
        combined_code += f"\n{'='*70}\n"
        combined_code += f"Arquivo: {filename}\n"
        combined_code += f"{'='*70}\n"
        combined_code += remove_comments(content)
        combined_code += "\n"
    
    # Limitar o tamanho do código se necessário (evitar tokens muito grandes)
    MAX_CHARS = 50000  # Ajuste conforme necessário
    if len(combined_code) > MAX_CHARS:
        combined_code = combined_code[:MAX_CHARS] + "\n\n[... código truncado devido ao tamanho ...]"
    
    system_message = (
        "You are a security researcher specialized in detecting security vulnerabilities.\n"
        "Analyze the provided code files and determine if they contain any security vulnerabilities.\n"
        "Provide the answer ONLY in the following format:\n\n"
        "vulnerability: <YES or NO> | vulnerability type: <type or N/A> | vulnerability name: <name or N/A> | explanation: <explanation for the prediction>.\n"
        "Do not include anything else in the response."
    )
    
    user_message = (
        f"Analyze the following code files from CVE: {cve_name} (Commit: {commit_hash[:8]}):\n\n"
        f"{combined_code}\n\n"
        "Is this code subject to any security vulnerability?\n"
        "Answer:"
    )
    
    try:
        chat_completion = client.chat.completions.create(
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message},
            ],
            model=model_name,
            temperature=0,
        )
        
        analysis_result = chat_completion.choices[0].message.content.strip()
        
        # Parse da resposta
        parsed_result = parse_llm_response(analysis_result)
        return parsed_result
        
    except Exception as e:
        return {
            "vulnerability": "ERROR",
            "vulnerability_type": "N/A",
            "vulnerability_name": "N/A",
            "explanation": f"Erro ao analisar com LLM: {str(e)}"
        }


def parse_llm_response(response):
    """
    Parse da resposta da LLM no formato esperado
    """
    result = {
        "vulnerability": "N/A",
        "vulnerability_type": "N/A",
        "vulnerability_name": "N/A",
        "explanation": response
    }
    
    try:
        # Tentar extrair os campos do formato especificado
        parts = response.split('|')
        for part in parts:
            part = part.strip()
            if part.lower().startswith('vulnerability:'):
                result["vulnerability"] = part.split(':', 1)[1].strip()
            elif part.lower().startswith('vulnerability type:'):
                result["vulnerability_type"] = part.split(':', 1)[1].strip()
            elif part.lower().startswith('vulnerability name:'):
                result["vulnerability_name"] = part.split(':', 1)[1].strip()
            elif part.lower().startswith('explanation:'):
                result["explanation"] = part.split(':', 1)[1].strip()
    except:
        pass
    
    return result


def process_all_cves(cves_folder, models_to_use, output_excel):
    """
    Processa todos os CVEs na pasta CVEs e gera um relatório Excel
    """
    all_results = []
    
    # Verificar se a pasta CVEs existe
    if not os.path.exists(cves_folder):
        print(f"[ERRO] Pasta '{cves_folder}' não encontrada!")
        return
    
    # Listar todas as pastas de CVEs
    cve_folders = [f for f in os.listdir(cves_folder) if os.path.isdir(os.path.join(cves_folder, f))]
    
    print(f"\n{'='*70}")
    print(f"  ANÁLISE DE CVEs COM LLMs")
    print(f"{'='*70}\n")
    print(f"Total de CVEs encontrados: {len(cve_folders)}")
    print(f"Modelos a serem utilizados: {len(models_to_use)}")
    print(f"{'='*70}\n")
    
    total_analyses = 0
    
    for cve_idx, cve_folder_name in enumerate(sorted(cve_folders), 1):
        cve_path = os.path.join(cves_folder, cve_folder_name)
        
        print(f"[{cve_idx}/{len(cve_folders)}] Processando CVE: {cve_folder_name}")
        
        # Listar todas as pastas de commits (hashes)
        commit_folders = [f for f in os.listdir(cve_path) 
                         if os.path.isdir(os.path.join(cve_path, f)) and len(f) == 40]
        
        if not commit_folders:
            print(f"   ℹ Nenhum commit encontrado para {cve_folder_name}")
            continue
        
        print(f"   🔍 {len(commit_folders)} commit(s) encontrado(s)")
        
        for commit_idx, commit_hash in enumerate(sorted(commit_folders), 1):
            commit_path = os.path.join(cve_path, commit_hash)
            
            print(f"      [{commit_idx}/{len(commit_folders)}] Commit: {commit_hash[:8]}...")
            
            # Ler todos os arquivos de código do commit
            code_files = read_all_code_files(commit_path)
            
            if not code_files:
                print(f"         ⚠ Nenhum arquivo de código encontrado")
                continue
            
            print(f"         📄 {len(code_files)} arquivo(s) de código encontrado(s)")
            
            # Analisar com cada modelo
            for model_idx, model_name in enumerate(models_to_use, 1):
                print(f"         [{model_idx}/{len(models_to_use)}] Analisando com {model_name}...", end=" ")
                
                try:
                    analysis = analyze_commit_with_llm(
                        cve_folder_name, 
                        commit_hash, 
                        code_files, 
                        model_name
                    )
                    
                    # Adicionar aos resultados
                    all_results.append({
                        "CVE": cve_folder_name,
                        "Commit Hash": commit_hash,
                        "Modelo": model_name,
                        "Vulnerabilidade Detectada": analysis["vulnerability"],
                        "Tipo de Vulnerabilidade": analysis["vulnerability_type"],
                        "Nome da Vulnerabilidade": analysis["vulnerability_name"],
                        "Explicação": analysis["explanation"],
                        "Quantidade de Arquivos": len(code_files),
                        "Arquivos Analisados": ", ".join(code_files.keys())
                    })
                    
                    total_analyses += 1
                    print("✓")
                    
                except Exception as e:
                    print(f"✗ Erro: {e}")
                    all_results.append({
                        "CVE": cve_folder_name,
                        "Commit Hash": commit_hash,
                        "Modelo": model_name,
                        "Vulnerabilidade Detectada": "ERROR",
                        "Tipo de Vulnerabilidade": "N/A",
                        "Nome da Vulnerabilidade": "N/A",
                        "Explicação": f"Erro: {str(e)}",
                        "Quantidade de Arquivos": len(code_files),
                        "Arquivos Analisados": ", ".join(code_files.keys())
                    })
        
        print()
    
    # Gerar relatório Excel
    print(f"\n{'='*70}")
    print(f"Gerando relatório Excel...")
    print(f"{'='*70}\n")
    
    generate_excel_report(all_results, output_excel, models_to_use)
    
    print(f"\n{'='*70}")
    print(f"  RESUMO FINAL")
    print(f"{'='*70}")
    print(f"  • CVEs processados: {len(cve_folders)}")
    print(f"  • Total de análises: {total_analyses}")
    print(f"  • Modelos utilizados: {len(models_to_use)}")
    print(f"  • Relatório salvo em: {output_excel}")
    print(f"{'='*70}\n")


def generate_excel_report(results, output_path, models_used):
    """
    Gera um relatório Excel com os resultados das análises
    """
    def _sanitize_sheet_title(name, existing_titles):
        invalid_chars = r'\\/*:?[]'
        for ch in invalid_chars:
            name = name.replace(ch, '-')
        name = name.strip()
        max_len = 31
        base = name[:max_len]
        candidate = base
        i = 1
        while candidate in existing_titles:
            suffix = f"-{i}"
            cut = max_len - len(suffix)
            candidate = base[:cut] + suffix
            i += 1
        return candidate

    # Estilos (reutilizáveis por sheet)
    title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=11, bold=True)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side,
                        top=thin_border_side, bottom=thin_border_side)
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    headers = [
        "CVE",
        "Commit Hash",
        "Modelo LLM",
        "Vulnerabilidade?",
        "Tipo",
        "Nome",
        "Explicação",
        "Qtd Arquivos",
        "Arquivos Analisados"
    ]

    workbook = Workbook()
    # remover sheet padrão criado pelo Workbook() para criar as planilhas ordenadas pelos modelos
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    existing_titles = set()

    for model in models_used:
        safe_title = _sanitize_sheet_title(model, existing_titles)
        existing_titles.add(safe_title)
        sheet = workbook.create_sheet(title=safe_title)

        # Cabeçalho/Título por sheet
        sheet.merge_cells('A1:I1')
        title_cell = sheet['A1']
        title_cell.value = f"Relatório - {model} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = title_fill
        sheet.row_dimensions[1].height = 24

        # Cabeçalhos das colunas
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num, value=header_title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        sheet.row_dimensions[2].height = 22

        # Filtrar resultados para o modelo atual
        model_results = [r for r in results if r.get('Modelo') == model]

        current_row = 3
        if not model_results:
            sheet.cell(row=current_row, column=1, value="Nenhum resultado encontrado para este modelo.")
            sheet.merge_cells(start_row=current_row, start_column=1,
                             end_row=current_row, end_column=len(headers))
        else:
            for result in model_results:
                data = [
                    result.get("CVE"),
                    result.get("Commit Hash", "")[:8],
                    result.get("Modelo"),
                    result.get("Vulnerabilidade Detectada"),
                    result.get("Tipo de Vulnerabilidade"),
                    result.get("Nome da Vulnerabilidade"),
                    result.get("Explicação"),
                    result.get("Quantidade de Arquivos"),
                    result.get("Arquivos Analisados")
                ]

                for col_num, cell_value in enumerate(data, 1):
                    cell = sheet.cell(row=current_row, column=col_num, value=cell_value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

                current_row += 1

        # Ajustar larguras das colunas por sheet
        sheet.column_dimensions[get_column_letter(1)].width = 20  # CVE
        sheet.column_dimensions[get_column_letter(2)].width = 12  # Commit Hash
        sheet.column_dimensions[get_column_letter(3)].width = 35  # Modelo
        sheet.column_dimensions[get_column_letter(4)].width = 15  # Vulnerabilidade?
        sheet.column_dimensions[get_column_letter(5)].width = 20  # Tipo
        sheet.column_dimensions[get_column_letter(6)].width = 25  # Nome
        sheet.column_dimensions[get_column_letter(7)].width = 60  # Explicação
        sheet.column_dimensions[get_column_letter(8)].width = 12  # Qtd Arquivos
        sheet.column_dimensions[get_column_letter(9)].width = 50  # Arquivos

    # Salvar
    try:
        workbook.save(output_path)
        print(f"✓ Relatório Excel gerado com sucesso!")
    except Exception as e:
        print(f"✗ Erro ao salvar Excel: {e}")


if __name__ == "__main__":
    # Configurações
    CVES_FOLDER = "CVEs"
    OUTPUT_EXCEL = "Relatorio_Analise_CVEs_LLM.xlsx"
    
    # Modelos a serem utilizados
    GROQ_MODELS = [
        "moonshotai/kimi-k2-instruct-0905",
        "qwen/qwen3-32b",
        "gemma2-9b-it",
        "meta-llama/Llama-Guard-4-12B",
        "llama-3.3-70b-versatile",
        "llama-3.1-8b-instant",
        "meta-llama/llama-4-maverick-17b-128e-instruct",
        "meta-llama/llama-4-scout-17b-16e-instruct",
    ]
    
    # Processar todos os CVEs
    process_all_cves(CVES_FOLDER, GROQ_MODELS, OUTPUT_EXCEL)
